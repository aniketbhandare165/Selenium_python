import openpyxl
book = openpyxl.load_workbook("D:\\New_project_3\\Selenium\\pythonSelenium\\PythonDemo.xlsx")
sheet = book.active
samp_dict = {}
cell =sheet.cell(row=1,column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "Rahul"
print(sheet.cell(row=2, column=2).value)

print(sheet.max_column)
print(sheet.max_row)
print((sheet['A2'].value))

for i in range(1,sheet.max_row+1): #to get rows
    if sheet.cell(row=i,column=1).value == "Test case2":
        for j in range(2,sheet.max_column+1):  # to get columns
            samp_dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value
            # print(sheet.cell(row=i,column=j).value)

print(samp_dict) 