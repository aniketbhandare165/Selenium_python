import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

def update_excel_data(filepath, searchTerm, column, val):
    book = openpyxl.load_workbook(filepath)
    sheet = book.active
    Dict = {}
    for i in range(1,sheet.max_column+1):
        if sheet.cell(row=1,column=i).value == column:
            Dict["col"] = i
    for i in range(1,sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            if sheet.cell(row=i,column=j).value == searchTerm:
                Dict["row"]=i
    sheet.cell(row=Dict["row"], column=Dict["col"]).value = val
    book.save(file_path)

file_path = "C:\\Users\\anike\\Downloads\\download.xlsx"
fruit_name = "Apple"
newvalue="999"
driver = webdriver.Chrome()
driver.get("https://rahulshettyacademy.com/upload-download-test/")
driver.implicitly_wait(3)
driver.find_element(By.ID,"downloadButton").click()
time.sleep(6)
# Edit excel with updated value
update_excel_data(file_path, fruit_name, "price", newvalue)

#upload
file_input = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
file_input.send_keys(file_path)
toast_locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait = WebDriverWait(driver, 10).until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

price_column = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id ='cell-"+price_column+"-undefined']").text
assert actual_price == newvalue
print(actual_price)